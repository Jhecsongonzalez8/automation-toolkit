import pywhatkit as kit
from openpyxl import load_workbook
import time
import customtkinter as ctk


def executar():
    planilha = load_workbook("clientes.xlsx")
    aba = planilha.active

    for linha in aba.iter_rows(min_row=2, values_only=True):
        nome, numero, vencimento = linha[:3]
        data = vencimento.strftime("%d/%m/%Y")

        mensagem = f"Olá {nome}, tudo bem? Sua dívida vence em {data}. Faça o pagamento no link: pagamento.com"

        kit.sendwhatmsg_instantly(
            f"+55{numero}",
            mensagem,
            wait_time=15,
            tab_close=True
        )

        time.sleep(20)
pass


#    aparecncia
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

#app(desenho)
app= ctk.CTk()
app.geometry('500x300')
app.title('--BOT DE MENSAGEMS--')
#titulo
titulo =ctk.CTkLabel(app,text='Painel de automação', font=("Arial", 22))
titulo.pack(pady=10)
#botao
botao = ctk.CTkButton(app, text="Executar", command=executar)
botao.pack(pady=20)
app.mainloop()

   